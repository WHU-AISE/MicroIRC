#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@author: zhuyuhan2333
"""

from os import error
import os
import pandas as pd
from util import formalizeDataFrame
from openpyxl import load_workbook
from openpyxl import Workbook
from graphsage_metric.time import Time
import time
from MicroIRC import dfTimelimit

def find_last(string,str):
    last_position=-1
    while True:
        position=string.find(str,last_position+1)
        if position==-1:
            return last_position
        last_position=position

def parse_svc_data(folder):
    minute = 10
    label_list = ['2022-7-23 ']
    label_file_name = folder + '/' + 'label-20220723' + '.csv'
    label_data = pd.read_csv(label_file_name, encoding='utf-8')
    root_causes = label_data['cmdb_id']
    time_list = []

    for row in label_data.itertuples():
        root_cause = row[3]
        root_cause_level = row[2]
        real_time = label_list[0] + row[1]
        real_timestamp = int(time.mktime(time.strptime(real_time, "%Y-%m-%d %H:%M:%S")))
        begin_timestamp = real_timestamp - 60 * minute;
        end_timestamp = real_timestamp + 60 * minute;
        t = Time(begin_timestamp, end_timestamp, root_cause, root_cause_level)
        time_list.append(t)
    metric_file_name = folder + '/' + 'metric.csv'
    row_data = pd.read_csv(metric_file_name)
    for t in time_list:
        root_cause = t.root_cause
        root_cause_level = t.root_cause_level
        begin_timestamp = t.begin
        end_timestamp = t.end
        metric_source_data = dfTimelimit(row_data, begin_timestamp, end_timestamp)
        # metric_source_data = metric_source_data.set_index('timestamp')

        headers = metric_source_data.columns
        data_prefix = metric_source_data.iloc[:,[0]]
        # data_suffix = metric_source_data.iloc[:,[-3, -2, -1]]

        # node data
        # node_file_name = folder + '/' + 'node.csv'
        # node_source_data = pd.read_csv(node_file_name)
        # for head in node_source_data.columns:
        #     if 'node' not in head:
        #         node_source_data = node_source_data.drop([head], axis=1)

        data_map = {}
        data_max_diff = {}
        for head in headers:
            if 'timestamp' in head: continue
            # temp_name = head[0:head.find('&')]
            temp_name = head[0:head.find('-')]
            data_f = metric_source_data[head].to_frame()
            data = formalizeDataFrame(data_f)
            diff = max(data) - min(data)
            try:
                diff_now = data_max_diff[temp_name]
            except:
                diff_now = 0
            if diff > diff_now:
                data_max_diff[temp_name] = diff
                data_map[temp_name] = data.squeeze()
        w = Workbook()
        ws = w.create_sheet()
        r = 1
        for key in data_map:
            c = 1
            ws.cell(row=r, column=c).value = key
            for i in range(len(data_map[key])):
                c += 1
                ws.cell(row=r, column=c).value = data_map[key][i]
            r += 1
        fileName = folder + '/rawdata-' + root_cause + '.xlsx'
        if not os.path.exists(os.path.dirname(fileName)):
            os.makedirs(os.path.dirname(fileName))
        w.save(fileName)

def parse_access_data(folder):
    calls = {}
    svc_map = {
                0:'adservice',
                1:'cartservice',
                2:'checkoutservice',
                3:'currencyservice',
                4:'emailservice',
                5:'frontend',
                6:'paymentservice',
                7:'productcatalogservice',
                8:'recommendationservice',
                9:'shippingservice'
                }
    call = folder + '/' + 'call.csv'
    call_data = pd.read_csv(call)
    for head in call_data.columns:
        if 'Unknown' not in head and 'redis' not in head and head != 'timestamp' and 'Unnamed' not in head and 'node' not in head and 'tcp' not in head:
            try:
                caller = calls[head[0:head.find('_')]]
            except:
                caller = set([])
            caller.add(head[head.find('_')+1:head.find('&')])
            calls[head[0:head.find('_')]] = caller
    w = Workbook()
    ws = w.create_sheet()
    for i in range(10):
        for j in range(10):
            try:
                if svc_map[j] in calls[svc_map[i]]:
                    ws.cell(row=i + 1, column=j + 1).value = 1
                else:
                    ws.cell(row=i + 1, column=j + 1).value = 0
            except:
                ws.cell(row=i + 1, column=j + 1).value = 0
    fileName = folder + '/access.xlsx'
    if not os.path.exists(os.path.dirname(fileName)):
        os.makedirs(os.path.dirname(fileName))
    w.save(fileName)


if __name__ == '__main__':
    folder = './20220723'
    parse_svc_data(folder)
    # parse_access_data(folder)
    print('yes')